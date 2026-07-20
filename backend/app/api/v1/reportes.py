from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, and_
from app.core.database import get_db
from app.api.deps import get_current_admin
from app.models.usuario import Usuario
from app.models.conductor import Conductor
from app.models.solicitud_viaje import SolicitudViaje
from app.models.pago import Pago
from datetime import datetime, timedelta
import io
import openpyxl
from openpyxl.styles import Font

router = APIRouter()


@router.get("/dashboard")
async def obtener_metricas_dashboard(
    periodo: str = Query(
        "7dias", description="Periodo de tiempo: hoy, 7dias, 30dias, año"
    ),
    current_user: Usuario = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db),
):
    ahora = datetime.now()
    if periodo == "hoy":
        fecha_limite = ahora.replace(hour=0, minute=0, second=0, microsecond=0)
    elif periodo == "7dias":
        fecha_limite = ahora - timedelta(days=7)
    elif periodo == "30dias":
        fecha_limite = ahora - timedelta(days=30)
    else:  # año
        fecha_limite = ahora - timedelta(days=365)

    # 1. Total usuarios activos
    res_usr = await db.execute(
        select(func.count(Usuario.id)).where(Usuario.estado == "activo")
    )
    usuarios_activos = res_usr.scalar() or 0

    # 2. Viajes completados hoy/en el periodo
    res_viajes = await db.execute(
        select(func.count(SolicitudViaje.id)).where(
            and_(
                SolicitudViaje.estado == "completada",
                SolicitudViaje.creado_en >= fecha_limite,
            )
        )
    )
    viajes_completados = res_viajes.scalar() or 0

    # 3. KMs Ahorrados (distancia total recorrida en viajes compartidos)
    res_kms = await db.execute(
        select(func.sum(SolicitudViaje.distancia_viaje_km)).where(
            and_(
                SolicitudViaje.estado == "completada",
                SolicitudViaje.creado_en >= fecha_limite,
            )
        )
    )
    kms_ahorrados = float(res_kms.scalar() or 0.0)

    # 4. Ahorro monetario estimado o comisiones cobradas
    # El ahorro para el pasajero es significativo respecto a un taxi.
    # El dinero generado en comisiones por la app:
    res_comision = await db.execute(
        select(func.sum(Pago.monto_comision_app_bs)).where(
            and_(Pago.estado == "completado", Pago.creado_en >= fecha_limite)
        )
    )
    dinero_generado_comisiones = float(res_comision.scalar() or 0.0)

    # Gráfico de evolución de viajes completados por día
    # Agrupamos por fecha truncada a día
    res_grafico = await db.execute(
        select(
            func.date(SolicitudViaje.creado_en).label("fecha"),
            func.count(SolicitudViaje.id).label("cantidad"),
        )
        .where(
            and_(
                SolicitudViaje.estado == "completada",
                SolicitudViaje.creado_en >= fecha_limite,
            )
        )
        .group_by(func.date(SolicitudViaje.creado_en))
        .order_by("fecha")
    )
    filas_grafico = res_grafico.mappings().all()

    evolucion_viajes = []
    for f in filas_grafico:
        evolucion_viajes.append(
            {"fecha": f["fecha"].isoformat(), "cantidad": f["cantidad"]}
        )

    # Gráfico de nuevos usuarios por día en el periodo
    res_usuarios_grafico = await db.execute(
        select(
            func.date(Usuario.creado_en).label("fecha"),
            func.count(Usuario.id).label("cantidad"),
        )
        .where(and_(Usuario.estado == "activo", Usuario.creado_en >= fecha_limite))
        .group_by(func.date(Usuario.creado_en))
        .order_by("fecha")
    )
    filas_usuarios = res_usuarios_grafico.mappings().all()

    evolucion_usuarios = []
    for f in filas_usuarios:
        evolucion_usuarios.append(
            {"fecha": f["fecha"].isoformat(), "cantidad": f["cantidad"]}
        )

    return {
        "usuarios_activos": usuarios_activos,
        "viajes_hoy": viajes_completados,
        "kms_ahorrados": round(kms_ahorrados, 1),
        "comisiones_ganadas_bs": round(dinero_generado_comisiones, 2),
        "grafico_viajes": evolucion_viajes,
        "grafico_usuarios": evolucion_usuarios,
    }


@router.get("/morosos")
async def obtener_conductores_morosos(
    current_user: Usuario = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db),
):
    # Buscamos conductores que tengan comisiones pendientes mayores a 0
    result = await db.execute(
        select(Conductor)
        .where(Conductor.comisiones_pendientes_bs > 0)
        .order_by(Conductor.comisiones_pendientes_bs.desc())
    )
    conductores = result.scalars().all()

    morosos = []
    for c in conductores:
        dias_deuda = 0
        if c.fecha_inicio_deuda:
            dias_deuda = (datetime.now() - c.fecha_inicio_deuda).days

        morosos.append(
            {
                "conductor_id": str(c.id),
                "nombre": f"{c.usuario.nombre} {c.usuario.apellido}",
                "email": c.usuario.email,
                "telefono": c.usuario.telefono,
                "comisiones_pendientes_bs": float(c.comisiones_pendientes_bs),
                "fecha_inicio_deuda": c.fecha_inicio_deuda.isoformat()
                if c.fecha_inicio_deuda
                else None,
                "dias_deuda": dias_deuda,
                "cuenta_congelada": c.cuenta_congelada,
            }
        )
    return morosos


@router.get("/exportar-excel")
async def exportar_excel(
    periodo: str = Query(
        "7dias", description="Periodo de tiempo: hoy, 7dias, 30dias, año"
    ),
    current_user: Usuario = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db),
):
    ahora = datetime.now()
    if periodo == "hoy":
        fecha_limite = ahora.replace(hour=0, minute=0, second=0, microsecond=0)
    elif periodo == "7dias":
        fecha_limite = ahora - timedelta(days=7)
    elif periodo == "30dias":
        fecha_limite = ahora - timedelta(days=30)
    else:  # año
        fecha_limite = ahora - timedelta(days=365)

    # ── KPIs para hoja Resumen ──────────────────────────────────────────────
    res_usr = await db.execute(
        select(func.count(Usuario.id)).where(Usuario.estado == "activo")
    )
    usuarios_activos = res_usr.scalar() or 0

    res_viajes = await db.execute(
        select(func.count(SolicitudViaje.id)).where(
            and_(
                SolicitudViaje.estado == "completada",
                SolicitudViaje.creado_en >= fecha_limite,
            )
        )
    )
    viajes_completados = res_viajes.scalar() or 0

    res_kms = await db.execute(
        select(func.sum(SolicitudViaje.distancia_viaje_km)).where(
            and_(
                SolicitudViaje.estado == "completada",
                SolicitudViaje.creado_en >= fecha_limite,
            )
        )
    )
    kms_ahorrados = float(res_kms.scalar() or 0.0)

    res_comision = await db.execute(
        select(func.sum(Pago.monto_comision_app_bs)).where(
            and_(Pago.estado == "completado", Pago.creado_en >= fecha_limite)
        )
    )
    comisiones_ganadas = float(res_comision.scalar() or 0.0)

    # ── Conductores morosos ─────────────────────────────────────────────────
    res_morosos = await db.execute(
        select(Conductor)
        .where(Conductor.comisiones_pendientes_bs > 0)
        .order_by(Conductor.comisiones_pendientes_bs.desc())
    )
    conductores_morosos = res_morosos.scalars().all()

    # ── Viajes por día ──────────────────────────────────────────────────────
    res_grafico = await db.execute(
        select(
            func.date(SolicitudViaje.creado_en).label("fecha"),
            func.count(SolicitudViaje.id).label("cantidad"),
        )
        .where(
            and_(
                SolicitudViaje.estado == "completada",
                SolicitudViaje.creado_en >= fecha_limite,
            )
        )
        .group_by(func.date(SolicitudViaje.creado_en))
        .order_by("fecha")
    )
    filas_grafico = res_grafico.mappings().all()

    # ── Construir Excel en memoria ──────────────────────────────────────────
    wb = openpyxl.Workbook()
    bold = Font(bold=True)

    # Hoja 1: Resumen
    ws_resumen = wb.active
    ws_resumen.title = "Resumen"
    cabeceras_resumen = ["Métrica", "Valor"]
    ws_resumen.append(cabeceras_resumen)
    for cell in ws_resumen[1]:
        cell.font = bold
    ws_resumen.append(["Usuarios Activos", usuarios_activos])
    ws_resumen.append(["Viajes Completados", viajes_completados])
    ws_resumen.append(["Comisiones Ganadas (Bs.)", round(comisiones_ganadas, 2)])
    ws_resumen.append(["KMs Ahorrados", round(kms_ahorrados, 1)])

    # Hoja 2: Conductores Morosos
    ws_morosos = wb.create_sheet("Conductores Morosos")
    cabeceras_morosos = [
        "Nombre",
        "Apellido",
        "Email",
        "Comisiones Pendientes (Bs.)",
        "Cuenta Congelada",
    ]
    ws_morosos.append(cabeceras_morosos)
    for cell in ws_morosos[1]:
        cell.font = bold
    for c in conductores_morosos:
        ws_morosos.append(
            [
                c.usuario.nombre,
                c.usuario.apellido,
                c.usuario.email,
                float(c.comisiones_pendientes_bs),
                "Sí" if c.cuenta_congelada else "No",
            ]
        )

    # Hoja 3: Viajes por Día
    ws_viajes = wb.create_sheet("Viajes por Dia")
    cabeceras_viajes = ["Fecha", "Cantidad"]
    ws_viajes.append(cabeceras_viajes)
    for cell in ws_viajes[1]:
        cell.font = bold
    for fila in filas_grafico:
        ws_viajes.append(
            [
                fila["fecha"].isoformat()
                if hasattr(fila["fecha"], "isoformat")
                else str(fila["fecha"]),
                fila["cantidad"],
            ]
        )

    # Guardar en memoria y retornar
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    headers = {"Content-Disposition": "attachment; filename=reporte_apuradito.xlsx"}
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )
